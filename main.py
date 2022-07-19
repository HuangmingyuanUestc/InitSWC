# coding=utf-8
import copy
import glob
import os
from shutil import copyfile
import openpyxl as op
import requests
import wget
import xlrd
import pdfkit
from bs4 import BeautifulSoup

# 配置：
# 1. 指定人名（要在表内）：
TARGET_HUMAN_NAME = 'huangmingyuan'
# 2. 下载：https://wkhtmltopdf.org/downloads.html，并配置wkhtmltopdf的下载路径
WKHTML_PATH = r'C:\Develop\Tools\wkhtmltopdf\bin\wkhtmltopdf.exe'
# 3. 配置是否下载报告和代码，以及是否需要初始化分析表
NEED_DOWNLOAD_REPORT = False
NEED_DOWNLOAD_CODE = False
NEED_DOWNLOAD_SHEET = True
NEED_REFRESH_SHEET = False
NEED_OPENZEPPELIN_FINDING = False


def get_form_data(dir_case, sheetnum):
    data = xlrd.open_workbook(dir_case)
    table = data.sheets()[sheetnum]
    nor = table.nrows
    nol = table.ncols
    dict = {}
    for i in range(1, nor):
        for j in range(nol):
            title = table.cell_value(0, j)
            value = table.cell_value(i, j)
            # print value
            dict[title] = value
        yield dict


def save_to_pdf(url, target, wkhtmltopdf_path):
    try:
        # 本来直接调用pdfkid的from方法就可以了，但是由于我们的wkhtmltopdf安装包有点问题，一直没法搜到，所以只能用本办法，直接配置了wk的地址
        # wkhtmltopdf下载链接： https://wkhtmltopdf.org/downloads.html
        config = pdfkit.configuration(wkhtmltopdf=wkhtmltopdf_path)
        pdfkit.from_url(url, target, configuration=config)
    except:
        pass


def save_to_zip(code_url, code_path):
    try:
        wget.download(code_url, code_path)
    except:
        pass


def get_finding(report_url, sheet_path):
    try:
        req = requests.get(report_url)
        req.encoding = "utf-8"
        html = req.text
        soup = BeautifulSoup(req.text, features="html.parser")
        li_all = soup.find_all(name='h3')
        li_h4 = soup.find_all(name='h4')
        li_new = []

        for li_item in li_all:
            # if li_item.text[0] == "[":
            li_new.append(li_item.text)
        for h4_item in li_h4:
            li_new.append(h4_item.text)

        print(sheet_path)
        print(li_new)
        bg = op.load_workbook(sheet_path)  # 应先将excel文件放入到工作目录下
        sheet = bg["Sheet1"]  # “Sheet1”表示将数据写入到excel文件的sheet1下
        for i in range(1, len(li_new) + 1):
            sheet.cell(i+1, 4, li_new[i - 1])  # sheet.cell(1,1,num_list[0])表示将num_list列表的第0个数据1写入到excel表格的第一行第一列
        bg.save(sheet_path)  # 对文件进行保存

    except:
        pass


if __name__ == '__main__':
    # 1 初始化数据到程序
    full_data_list = []
    target_data_list = []
    # 1.1 导入表格数据
    for item in get_form_data(r'input.xlsx', 0):
        full_data_list.append(copy.deepcopy(item))
    # 1.2 按人名筛选表格数据
    for data_item in full_data_list:
        if data_item['人员'] == TARGET_HUMAN_NAME or data_item['人员'] == TARGET_HUMAN_NAME:
            target_data_list.append(copy.deepcopy(data_item))
    print(target_data_list)

    for data_item in target_data_list:
        # 2.0 初始化路径
        file_path = 'output/' + str(data_item["公司"]) + '-' + str(data_item["dapp"])
        file_path = file_path.strip()
        report_path = file_path + "/report.pdf"
        code_path = file_path + "/code.zip"
        code_url = data_item["对应的仓库链接（没有的话写无）"].replace('/tree/', '/archive/', 1) + '.zip'
        sheet_path = file_path + "/analysis.xlsx"
        example_sheet_path = 'analysis-example.xlsx'
        # 2.1 创建目录
        if not os.path.exists(file_path):
            os.makedirs(file_path)
        # 2.2 下载报告
        if NEED_DOWNLOAD_REPORT:
            if not os.path.exists(report_path):
                print('Download Report[' + report_path + ']: ' + data_item["报告链接"])
                save_to_pdf(data_item["报告链接"], report_path, WKHTML_PATH)
        # 2.3 下载代码
        if NEED_DOWNLOAD_CODE:
            if not glob.glob(file_path + "/*.zip"):
                print('Download Code[' + code_path + ']: ' + data_item["对应的仓库链接（没有的话写无）"])
                save_to_zip(code_url, code_path)
        # 2.4 初始化表格
        if NEED_DOWNLOAD_SHEET:
            if not os.path.exists(sheet_path):
                print('Copy Sheet[' + sheet_path + ']: ' + example_sheet_path)
                copyfile(example_sheet_path, sheet_path)
        # 2.4.X 辅助选项：刷新表格
        if NEED_REFRESH_SHEET:
            if os.path.exists(sheet_path):
                os.remove(sheet_path)
            if not os.path.exists(sheet_path):
                print('Copy Sheet[' + sheet_path + ']: ' + example_sheet_path)
                copyfile(example_sheet_path, sheet_path)
        # 2.5 初始化finding内容
        if NEED_OPENZEPPELIN_FINDING:
            if os.path.exists(sheet_path):
                get_finding(data_item["报告链接"], sheet_path)
